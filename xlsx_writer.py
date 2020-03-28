import openpyxl as xl
import os, re, time, xlsxwriter
from openpyxl.utils import get_column_letter
from progress.bar import Bar

"""
川 <--- 行

三 <--- 列
"""

wbook = xl.Workbook()
wsheet = wbook.active
wsheet['A1'] = 'ID'
wsheet['B1'] = '課程名稱'            # Course_Title
wsheet['C1'] = '授課教師'            # Course_Instructor
wsheet['D1'] = '開課系級(中)'        # Course_Class(Chinese)
wsheet['E1'] = '開課系級(英)'        # Course_Class(English)
wsheet['F1'] = '開課資料'            # Course_Details

wsheet['G1'] = '目標類型'            # Stats_ObjectiveMethods
wsheet['H1'] = '核心能力'            # Stats_CoreCompetences
wsheet['I1'] = '基本素養'            # Stats_EssentialVirtues 
wsheet['J1'] = '教學方法'            # Stats_TeachingMethods
wsheet['K1'] = '評量方式'            # Stats_Assessment

wsheet['L1'] = '出席率'              # Grading_Attendance
wsheet['M1'] = '平時評量'            # Grading_MarkofUsual 
wsheet['N1'] = '期中評量'            # Grading_MidtermExam 
wsheet['O1'] = '期末評量'            # Grading_FinalExam 
wsheet['P1'] = '其他'                # Grading_Other
wsheet['Q1'] = '其他<>'              # Grading_Other<>

directory = './xlsx'

start_time = time.time()
with Bar('Processing', max=4371) as bar:
    for filename in sorted(os.listdir(directory)):
        data = []
        pdfpath=os.path.join(directory, filename) # 將路徑變成 ./xlsx/xxxx.xlsx
        rbook = xl.load_workbook(pdfpath, data_only=True)
        rsheet = rbook.worksheets[0]

        
        # 將檔案名存為 ID，檔案名是取自：http://ap09.emis.tku.edu.tw/108_2/108_2_檔案名.PDF
        data.append(os.path.splitext(filename)[0])                      # ID

        # 「課程名稱」、「授課教師」都會在第一行
        for cell in rsheet['1']:
            if cell.value == "課程名稱" or cell.value == "Course Title":
                coordinate = "{column}{row}".format(column=get_column_letter(cell.column+1), row=cell.row)
                ######### quick and dirty fix #########
                if rsheet[coordinate].value == None:
                    coordinate = "{column}{row}".format(column=get_column_letter(cell.column+2), row=cell.row)
                data.append(str(rsheet[coordinate].value).replace('\n', "")) # 課程名稱
            
            if cell.value == "授課\n教師" or cell.value == "Instructor":
                coordinate = "{column}{row}".format(column=get_column_letter(cell.column+1), row=cell.row)
                ######### quick and dirty fix #########
                if rsheet[coordinate].value == None:
                    coordinate = "{column}{row}".format(column=get_column_letter(cell.column+2), row=cell.row)
                data.append(str(rsheet[coordinate].value).replace('\n', "")) # 授課教師

        # 英文版教學計畫表的「開課系級」、「開課系級(代碼)」、「開課資料」都會在第二行
        for cell in rsheet['2']:
            try:
                if cell.value == "Course Class":
                    coordinate = "{column}{row}".format(column=get_column_letter(cell.column+1), row=cell.row)
                    ######### quick and dirty fix #########
                    if rsheet[coordinate].value == None:
                        coordinate = "{column}{row}".format(column=get_column_letter(cell.column+2), row=cell.row)
                    data.append(str(rsheet[coordinate].value).splitlines()[0])    # 開課系級
                    data.append(str(rsheet[coordinate].value).splitlines()[1])    # 開課系級(代碼)
                if cell.value == "Details":
                    coordinate = "{column}{row}".format(column=get_column_letter(cell.column+1), row=cell.row)
                    data.append(str(rsheet[coordinate].value).replace('\n', "")) # 開課資料
            except IndexError as ex:
                template = "    File: {0} Arguments: {1!r}"
                message = template.format(os.path.splitext(filename)[0], ex.args)
                print(message)
                data.append("❌")
                continue

        # 中文版教學計畫表的「開課系級」、「開課系級(代碼)」、「開課資料」都會在第三行
        for cell in rsheet['3']:
            if cell.value == "開課系級":
                coordinate = "{column}{row}".format(column=get_column_letter(cell.column+1), row=cell.row)
                data.append(rsheet[coordinate].value)                   # 開課系級
                coordinate = "{column}{row}".format(column=get_column_letter(cell.column+1), row=cell.row+1)
                data.append(rsheet[coordinate].value)                   # 開課系級(代碼)
            if cell.value == "開課\n資料":
                coordinate = "{column}{row}".format(column=get_column_letter(cell.column+1), row=cell.row)
                data.append(str(rsheet[coordinate].value).replace('\n', "")) # 開課資料
        
        # 接下來的資料不一定會在同一行所以用逐行搜尋關鍵字的方式定位
        # 又因「目標類型」不只一個地方會出現，所以改搜「院、系(所) 核心能力」
        
        StatsStartRow = 0       ###########################
        StatsStartCol = 0       #                         #
        StatsEndRow = 0         #  請忽略我的義大利麵 code  #
        StatsEndCol = 0         #                         #
        StatsCount = 0          ###########################
        
        for row in rsheet.iter_rows():
            for entry in row:
                try:
                    if '教學目標之目標類型、核心能力、基本素養教學方法與評量方式' in entry.value or 'The correspondences of teaching objectives : core competences, essential virtues, teaching methods, and assessment' in entry.value:
                        StatsStartRow = entry.row
                        StatsStartCol = entry.column
                except (AttributeError, TypeError):
                    continue

        for row in rsheet.iter_rows():
            for entry in row:
                try:
                    if '授 課 進 度 表' in entry.value or 'Course Schedule' in entry.value:
                        StatsEndRow = entry.row
                        StatsEndCol = entry.column
                except (AttributeError, TypeError):
                    continue

        StatsCount = StatsEndRow - StatsStartRow - 1

        coordinate = '{column}{row}'.format(column = get_column_letter(StatsStartCol + 1), row = StatsStartRow + 2)
        data.append(rsheet[coordinate].value)                           # 目標類型
        coordinate = '{column}{row}'.format(column = get_column_letter(StatsStartCol + 2), row = StatsStartRow + 2)
        data.append(rsheet[coordinate].value)                           # 核心能力
        coordinate = '{column}{row}'.format(column = get_column_letter(StatsStartCol + 3), row = StatsStartRow + 2)
        data.append(rsheet[coordinate].value)                           # 基本素養
        coordinate = '{column}{row}'.format(column = get_column_letter(StatsStartCol + 4), row = StatsStartRow + 2)
        data.append(rsheet[coordinate].value)                           # 教學方法
        coordinate = '{column}{row}'.format(column = get_column_letter(StatsStartCol + 5), row = StatsStartRow + 2)
        data.append(rsheet[coordinate].value)                           # 評量方式

        for row in rsheet.iter_rows():
            for entry in row:
                try:
                    if '出席率' in entry.value:
                        attendance = re.search(r"(?<=出席率：).*?(?=%)", entry.value)   # 出席率
                        data.append(str(attendance.group(0)).strip())
                except (AttributeError, TypeError):
                    continue
                try:
                    if '平時評量' in entry.value:
                        attendance = re.search(r"(?<=平時評量：).*?(?=%)", entry.value) # 平時評量
                        data.append(attendance.group(0))
                except (AttributeError, TypeError):
                    continue
                try:
                    if '期中評量' in entry.value:
                        midterm = re.search(r"(?<=期中評量：).*?(?=%)", entry.value)    # 期中評量
                        data.append(midterm.group(0))
                except (AttributeError, TypeError):
                    continue
                try:
                    if '期末評量' in entry.value:
                        final = re.search(r"(?<=期末評量：).*?(?=%)", entry.value)      # 期末評量
                        data.append(final.group(0))
                except (AttributeError, TypeError):
                    continue
                try:
                    if '其他〈' in entry.value:
                        other = re.search(r"(?<=其他〈).*?(?=〉：)", entry.value)       # 其他
                        data.append(other.group(0))
                except (AttributeError, TypeError):
                    continue
                try:
                    if '〉：' in entry.value:
                        other_num = re.search(r"(?<=〉：).*?(?=%)", entry.value)        # 其他<>
                        data.append(other_num.group(0))
                except (AttributeError, TypeError):
                    continue
        wsheet.append(data)
        if StatsCount > 1:
            for i in range(1,StatsCount):
                data = [ '', '', '', '', '', '']
                coordinate = '{column}{row}'.format(column = get_column_letter(StatsStartCol + 1), row = StatsStartRow + 2 + i)
                data.append(rsheet[coordinate].value)                           # 目標類型
                coordinate = '{column}{row}'.format(column = get_column_letter(StatsStartCol + 2), row = StatsStartRow + 2 + i)
                data.append(rsheet[coordinate].value)                           # 核心能力
                coordinate = '{column}{row}'.format(column = get_column_letter(StatsStartCol + 3), row = StatsStartRow + 2 + i)
                data.append(rsheet[coordinate].value)                           # 基本素養
                coordinate = '{column}{row}'.format(column = get_column_letter(StatsStartCol + 4), row = StatsStartRow + 2 + i)
                data.append(rsheet[coordinate].value)                           # 教學方法
                coordinate = '{column}{row}'.format(column = get_column_letter(StatsStartCol + 5), row = StatsStartRow + 2 + i)
                data.append(rsheet[coordinate].value)                           # 評量方式
                wsheet.append(data)
        rbook.close()
        bar.next()
    wbook.save("results.xlsx")
    wbook.close()
print("--- %s seconds ---" % (time.time() - start_time))