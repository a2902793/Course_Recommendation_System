import openpyxl as xl
import os, re
from openpyxl.utils import get_column_letter

directory = './testset/xlsx'
for filename in sorted(os.listdir(directory)):
    pdfpath=os.path.join(directory, filename) #./testset/ + 0001.pdf
    wb = xl.load_workbook(pdfpath, data_only=True)
    sheet = wb.worksheets[0]
    #row_count = sheet.max_row
    #column_count = sheet.max_column
    #print("Number of Rows    = %d" %row_count)
    for cell in sheet['1']:
        if cell.value == "課程名稱":
            coordinate = "%s1" %get_column_letter(cell.column+1)
            print("科目 = %s" %sheet[coordinate].value)

    for row in sheet.iter_rows():
        for entry in row:
            try:
                if '出席率' in entry.value:
                    attendance = re.search(r"(?<=出席率：).*?(?=%)", entry.value)
                    print("出席率 = %s" %attendance.group(0).strip())
                    #print(entry.value)
            except (AttributeError, TypeError):
                continue
            try:
                if '平時評量' in entry.value:
                    attendance = re.search(r"(?<=平時評量：).*?(?=%)", entry.value)
                    print("平時評量 = %s" %attendance.group(0))
                    #print(entry.value)
            except (AttributeError, TypeError):
                continue
            try:
                if '期中評量' in entry.value:
                    midterm = re.search(r"(?<=期中評量：).*?(?=%)", entry.value)
                    print("期中評量 = %s" %midterm.group(0))
                    #print(entry.value)
            except (AttributeError, TypeError):
                continue
            try:
                if '期末評量' in entry.value:
                    final = re.search(r"(?<=期末評量：).*?(?=%)", entry.value)
                    print("期末評量 = %s" %final.group(0))
                    #print(entry.value)
            except (AttributeError, TypeError):
                continue
            try:
                if '其他〈' in entry.value:
                    other = re.search(r"(?<=其他〈).*?(?=〉：)", entry.value)
                    print("其他〈〉 = %s" %other.group(0))
                    #print(entry.value)
            except (AttributeError, TypeError):
                continue
            try:
                if '〉：' in entry.value:
                    other_num = re.search(r"(?<=〉：).*?(?=%)", entry.value)
                    print("其他〈...〉 = %s" %other_num.group(0))
                    #print(entry.value)
            except (AttributeError, TypeError):
                continue
    print("\n")
    